"""
Creates the Excel input template: NVH_Report_Input.xlsx
Run this once to generate the input file, then fill it in.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

OUTPUT_PATH = os.path.join(os.path.dirname(__file__), "NVH_Report_Input.xlsx")

# Colours
BLUE_HEADER  = "1F4E79"
GOLD_HEADER  = "BF8F00"
LIGHT_BLUE   = "D6E4F0"
LIGHT_GOLD   = "FFF2CC"
WHITE        = "FFFFFF"
GREY_STEP7   = "E2EFDA"   # green-ish tint to flag Step-7 columns

def header_font(colour="FFFFFF"):
    return Font(name="Calibri", bold=True, color=colour, size=11)

def normal_font():
    return Font(name="Calibri", size=11)

def thin_border():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

def fill(hex_colour):
    return PatternFill("solid", fgColor=hex_colour)

wb = openpyxl.Workbook()

# ------------------------------------------------------------------ #
# Sheet 1: Campaign_Info
# ------------------------------------------------------------------ #
ws1 = wb.active
ws1.title = "Campaign_Info"

ws1.column_dimensions["A"].width = 28
ws1.column_dimensions["B"].width = 45

# Title row
ws1.merge_cells("A1:B1")
ws1["A1"] = "CAMPAIGN INFORMATION  (same for all samples in this campaign)"
ws1["A1"].font = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
ws1["A1"].fill = fill(BLUE_HEADER)
ws1["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[1].height = 22

fields = [
    ("Program Name",        "Stellantis CUSW 2-Speed PV PTU 131mm"),
    ("NVH Test Order",      "282084"),
    ("PTL Test Order",      "282083"),
    ("Test Dyno",           "RHTC AWD"),
    ("Pre-Test Date",       "2/10/2022"),
    ("Post-100% Date",      "3/2/2022"),
    ("Post-300% Date",      "n/a"),
    ("Design Level",        "PV (diff re-sourcing)"),
    ("Part Ratio",          "2.73"),
    ("Prop Info",           "PS1247"),
    ("Published By",        "C. Li"),
]

for i, (label, example) in enumerate(fields, start=2):
    ws1.cell(i, 1, label).font   = Font(name="Calibri", bold=True, size=11)
    ws1.cell(i, 1).fill          = fill(LIGHT_BLUE)
    ws1.cell(i, 1).border        = thin_border()
    ws1.cell(i, 1).alignment     = Alignment(vertical="center")

    ws1.cell(i, 2, example).font = normal_font()
    ws1.cell(i, 2).fill          = fill(WHITE)
    ws1.cell(i, 2).border        = thin_border()
    ws1.cell(i, 2).alignment     = Alignment(vertical="center")

# Note row
note_row = len(fields) + 3
ws1.merge_cells(f"A{note_row}:B{note_row}")
ws1.cell(note_row, 1, "Replace the example values above with your actual campaign data.").font = Font(
    name="Calibri", italic=True, color="595959", size=10)

# ------------------------------------------------------------------ #
# Sheet 2: Samples
# ------------------------------------------------------------------ #
ws2 = wb.create_sheet("Samples")

step2_cols = ["Sample #", "Part Number (P/N)", "Serial Number (S/N)",
              "Sample Number", "Published Date"]
step7_cols = ["Drive +450Nm (dB)", "Drive +75Nm (dB)", "Drive -75Nm (dB)",
              "Coast +450Nm (dB)", "Coast +75Nm (dB)", "Coast -75Nm (dB)"]

all_cols = step2_cols + step7_cols

# Title
ws2.merge_cells(f"A1:{get_column_letter(len(all_cols))}1")
ws2["A1"] = "SAMPLE DATA"
ws2["A1"].font = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
ws2["A1"].fill = fill(BLUE_HEADER)
ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 22

# Sub-headers (Step 2 / Step 7)
ws2.merge_cells(f"A2:{get_column_letter(len(step2_cols))}2")
ws2["A2"] = "STEP 2 — Fill in before generating reports"
ws2["A2"].font  = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
ws2["A2"].fill  = fill(BLUE_HEADER)
ws2["A2"].alignment = Alignment(horizontal="center", vertical="center")

c7_start = get_column_letter(len(step2_cols) + 1)
c7_end   = get_column_letter(len(all_cols))
ws2.merge_cells(f"{c7_start}2:{c7_end}2")
ws2[f"{c7_start}2"] = "STEP 7 — Fill in after engineering review, then run Update Results"
ws2[f"{c7_start}2"].font  = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
ws2[f"{c7_start}2"].fill  = fill(GOLD_HEADER)
ws2[f"{c7_start}2"].alignment = Alignment(horizontal="center", vertical="center")

# Column headers row 3
col_widths = [10, 22, 24, 16, 18, 18, 16, 16, 18, 16, 16]
for j, col_name in enumerate(all_cols, start=1):
    cell = ws2.cell(3, j, col_name)
    cell.font      = header_font("FFFFFF")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = thin_border()
    if j <= len(step2_cols):
        cell.fill = fill(BLUE_HEADER)
    else:
        cell.fill = fill(GOLD_HEADER)
    ws2.column_dimensions[get_column_letter(j)].width = col_widths[j - 1]

ws2.row_dimensions[3].height = 30

# Example data rows (2 example samples)
examples = [
    [1, "68333255AE", "T36P22024115524", "0002", "3/2/2022", "", "", "", "", "", ""],
    [2, "68333255AE", "T36P22024225525", "0003", "3/2/2022", "", "", "", "", "", ""],
]
for i, row_data in enumerate(examples, start=4):
    for j, val in enumerate(row_data, start=1):
        cell = ws2.cell(i, j, val if val != "" else None)
        cell.font   = normal_font()
        cell.border = thin_border()
        cell.alignment = Alignment(horizontal="center" if j in [1, 4] else "left",
                                   vertical="center")
        if j <= len(step2_cols):
            cell.fill = fill(LIGHT_BLUE)
        else:
            cell.fill = fill(GREY_STEP7)

# Add 8 more blank rows for additional samples
for i in range(6, 14):
    for j in range(1, len(all_cols) + 1):
        cell = ws2.cell(i, j)
        cell.border = thin_border()
        cell.fill   = fill(LIGHT_BLUE) if j <= len(step2_cols) else fill(GREY_STEP7)

# Note row
ws2.merge_cells(f"A15:{get_column_letter(len(all_cols))}15")
ws2.cell(15, 1,
    "Add one row per sample. For dB values: enter as a number only (e.g. -23). "
    "Negative = below target (likely PASS). Positive = above target (check carefully)."
).font = Font(name="Calibri", italic=True, color="595959", size=10)

# ------------------------------------------------------------------ #
# Sheet 3: Instructions
# ------------------------------------------------------------------ #
ws3 = wb.create_sheet("Instructions")
ws3.column_dimensions["A"].width = 80
ws3.row_dimensions[1].height = 22

instructions = [
    ("HOW TO USE THIS FILE", True, BLUE_HEADER, "FFFFFF"),
    ("", False, WHITE, "000000"),
    ("STEP 1 — Prepare your template (manual, one time per campaign)", True, LIGHT_BLUE, "000000"),
    ("  1. Open your previous report in PowerPoint.", False, WHITE, "000000"),
    ("  2. Replace all dB result values (e.g. -23dB) with XXdB.", False, WHITE, "000000"),
    ("  3. Leave everything else as-is.", False, WHITE, "000000"),
    ("  4. Save as:  report_template.pptx  in the working folder.", False, WHITE, "000000"),
    ("", False, WHITE, "000000"),
    ("STEP 2 — Generate reports for all samples", True, LIGHT_BLUE, "000000"),
    ("  1. Fill in the Campaign_Info sheet.", False, WHITE, "000000"),
    ("  2. Fill in the Samples sheet (Step 2 columns only).", False, WHITE, "000000"),
    ("  3. Double-click:  Step2_Generate_Reports.bat", False, WHITE, "000000"),
    ("  4. Reports appear in the  reports/  folder.", False, WHITE, "000000"),
    ("", False, WHITE, "000000"),
    ("STEPS 3-6 — Your engineering work (manual)", True, LIGHT_BLUE, "000000"),
    ("  Replace the ActivePictures, format them, and fill in the dB values.", False, WHITE, "000000"),
    ("", False, WHITE, "000000"),
    ("STEP 7 — Update Pass/Fail results", True, LIGHT_BLUE, "000000"),
    ("  1. Enter the dB values in the Step 7 columns of the Samples sheet.", False, WHITE, "000000"),
    ("  2. Double-click:  Step7_Update_Results.bat", False, WHITE, "000000"),
    ("  3. Reports are updated with dB values and Pass/Fail indicators.", False, WHITE, "000000"),
    ("  4. Manually verify and adjust if needed.", False, WHITE, "000000"),
    ("", False, WHITE, "000000"),
    ("PASS / FAIL RULES", True, LIGHT_BLUE, "000000"),
    ("  PASS             = dB value is <= +3dB above target  (all negative values are PASS)", False, WHITE, "000000"),
    ("  CONDITIONAL FAIL = dB value is between +3dB and +6dB above target", False, WHITE, "000000"),
    ("  ABSOLUTE FAIL    = dB value is more than +6dB above target", False, WHITE, "000000"),
]

for i, (text, bold, bg, fg) in enumerate(instructions, start=1):
    cell = ws3.cell(i, 1, text)
    cell.font      = Font(name="Calibri", bold=bold, color=fg, size=11)
    cell.fill      = fill(bg)
    cell.alignment = Alignment(vertical="center")
    if text:
        ws3.row_dimensions[i].height = 18

wb.save(OUTPUT_PATH)
print(f"Created: {OUTPUT_PATH}")
